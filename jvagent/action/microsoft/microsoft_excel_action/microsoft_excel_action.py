import io
import json
import logging
import re
from contextlib import asynccontextmanager
from typing import Annotated, Any, AsyncIterator, ClassVar, Dict, List, Optional
from urllib.parse import quote

import openpyxl
from jvspatial.core.annotations import attribute

from jvagent.action.google.google_sheets_action.google_sheets_action import (
    compose_a1_range,
    qualify_sheet_title,
    resolve_spreadsheet_id,
)
from jvagent.tooling.tool_decorator import tool

from ..microsoft_action import MicrosoftAction

logger = logging.getLogger(__name__)


def resolve_workbook_item_id(url_or_id: Optional[str]) -> str:
    """Resolve a OneDrive/SharePoint file URL or return a raw drive item id."""
    s = (url_or_id or "").strip()
    if not s:
        raise ValueError("Missing workbook id / URL")
    if "docs.google.com/spreadsheets" in s:
        return resolve_spreadsheet_id(s)
    m = re.search(r"/items/([^/?#]+)", s, re.I)
    if m:
        return m.group(1)
    return s


class MicrosoftExcelAction(MicrosoftAction):
    """Excel workbooks on OneDrive via Microsoft Graph workbook APIs."""

    worksheet_title: str = attribute(
        default="Sheet1",
        description="Default worksheet (tab) title when range has no sheet name",
    )
    spreadsheet_url: str = attribute(
        default="",
        description="Default workbook: OneDrive item id or sharing URL",
    )

    SCOPES: ClassVar[List[str]] = [
        "offline_access",
        "User.Read",
        "Files.ReadWrite.All",
    ]

    def _effective_worksheet_title(self, worksheet_title: Optional[str]) -> str:
        if worksheet_title is not None and str(worksheet_title).strip():
            return str(worksheet_title).strip()
        return self.worksheet_title

    def _resolve_workbook(self, spreadsheet_url_or_id: Optional[str] = None) -> str:
        if spreadsheet_url_or_id and str(spreadsheet_url_or_id).strip():
            return resolve_workbook_item_id(str(spreadsheet_url_or_id).strip())
        if self.spreadsheet_url and str(self.spreadsheet_url).strip():
            return resolve_workbook_item_id(str(self.spreadsheet_url).strip())
        raise ValueError(
            "Provide spreadsheet_url_or_id, or set spreadsheet_url on the MicrosoftExcelAction"
        )

    def _ws_seg(self, worksheet_title: str) -> str:
        q = qualify_sheet_title(worksheet_title or "")
        if q.startswith("'") and q.endswith("'"):
            inner = q[1:-1].replace("''", "'")
        else:
            inner = q
        inner = inner.replace("'", "''")
        return f"worksheets('{inner}')"

    def _local_a1(self, range_name: str, worksheet_title: Optional[str]) -> str:
        ws = self._effective_worksheet_title(worksheet_title)
        full = compose_a1_range(ws, range_name if range_name else None)
        if "!" in full:
            return full.split("!", 1)[1]
        return ""

    @asynccontextmanager
    async def _workbook_session(self, item_id: str) -> AsyncIterator[str]:
        sess = await self.graph_json(
            "POST",
            f"/me/drive/items/{item_id}/workbook/createSession",
            json_body={"persistChanges": True},
            ok=(200, 201),
        )
        sid = (sess or {}).get("id") or ""
        if not sid:
            raise RuntimeError("Could not create Excel workbook session")
        try:
            yield sid
        finally:
            try:
                await self.graph_request(
                    "POST",
                    f"/me/drive/items/{item_id}/workbook/closeSession",
                    json_body={"workbookSessionId": sid},
                )
            except Exception:
                logger.debug("closeSession failed", exc_info=True)

    async def read_spreadsheet(
        self,
        spreadsheet_url_or_id: Optional[str] = None,
        range_name: str = "",
        worksheet_title: Optional[str] = None,
    ) -> List[List[Any]]:
        item_id = self._resolve_workbook(spreadsheet_url_or_id)
        ws = self._effective_worksheet_title(worksheet_title)
        seg = self._ws_seg(ws)
        if not range_name.strip():
            path = f"/me/drive/items/{item_id}/workbook/{seg}/usedRange"
            data = await self.graph_json("GET", path, params={"valuesOnly": "true"})
        else:
            addr = self._local_a1(range_name, worksheet_title)
            path = f"/me/drive/items/{item_id}/workbook/{seg}/range(address='{addr}')"
            data = await self.graph_json("GET", path)

        vals = (data or {}).get("values") if isinstance(data, dict) else None
        return vals or []

    async def update_spreadsheet(
        self,
        spreadsheet_url_or_id: Optional[str] = None,
        range_name: str = "",
        values: Optional[List[List[Any]]] = None,
        value_input_option: str = "RAW",
        worksheet_title: Optional[str] = None,
    ) -> Dict[str, Any]:
        _ = value_input_option
        if values is None:
            raise ValueError("values is required")
        if not range_name.strip():
            raise ValueError("range_name is required for update")
        item_id = self._resolve_workbook(spreadsheet_url_or_id)
        ws = self._effective_worksheet_title(worksheet_title)
        seg = self._ws_seg(ws)
        addr = self._local_a1(range_name, worksheet_title)
        async with self._workbook_session(item_id) as session_id:
            path = f"/me/drive/items/{item_id}/workbook/{seg}/range(address='{addr}')"
            resp = await self.graph_request(
                "PATCH",
                path,
                json_body={"values": values},
                headers={"workbook-session-id": session_id},
            )
            if resp.status_code not in (200, 204):
                raise RuntimeError(
                    f"update range failed: {resp.status_code} {resp.text[:400]}"
                )
            return {"success": True} if resp.status_code == 204 else resp.json()

    async def append_spreadsheet(
        self,
        spreadsheet_url_or_id: Optional[str] = None,
        range_name: Optional[str] = None,
        values: Optional[List[List[Any]]] = None,
        value_input_option: str = "RAW",
        worksheet_title: Optional[str] = None,
    ) -> Dict[str, Any]:
        _ = value_input_option
        if values is None:
            raise ValueError("values is required")
        item_id = self._resolve_workbook(spreadsheet_url_or_id)
        ws_default = self._effective_worksheet_title(worksheet_title)
        existing = await self.read_spreadsheet(
            spreadsheet_url_or_id=item_id,
            range_name=range_name or "",
            worksheet_title=ws_default,
        )
        start_row = len(existing) + 1
        anchor = (range_name or "A").strip() or "A"
        col_part = "".join(ch for ch in anchor.split("!")[-1] if ch.isalpha()) or "A"
        addr = f"{col_part}{start_row}"
        return await self.update_spreadsheet(
            spreadsheet_url_or_id=item_id,
            range_name=addr,
            values=values,
            worksheet_title=ws_default,
        )

    async def batch_clear(
        self,
        spreadsheet_url_or_id: Optional[str] = None,
        ranges: Optional[List[str]] = None,
        worksheet_title: Optional[str] = None,
    ) -> Dict[str, Any]:
        if not ranges:
            raise ValueError("ranges is required")
        item_id = self._resolve_workbook(spreadsheet_url_or_id)
        ws = self._effective_worksheet_title(worksheet_title)
        seg = self._ws_seg(ws)
        cleared = 0
        async with self._workbook_session(item_id) as session_id:
            for r in ranges:
                r = (r or "").strip()
                if not r:
                    continue
                addr = self._local_a1(r, worksheet_title)
                path = (
                    f"/me/drive/items/{item_id}/workbook/{seg}/range(address='{addr}')"
                )
                await self.graph_request(
                    "PATCH",
                    path,
                    json_body={"values": []},
                    headers={"workbook-session-id": session_id},
                )
                cleared += 1
        return {"success": True, "clearedRanges": cleared}

    async def format_cells(self, *args: Any, **kwargs: Any) -> Dict[str, Any]:
        raise NotImplementedError(
            "Excel graph format_cells is not implemented; use update with raw values."
        )

    async def merge_cells(self, *args: Any, **kwargs: Any) -> Dict[str, Any]:
        raise NotImplementedError("Excel graph merge_cells is not implemented.")

    async def unmerge_cells(self, *args: Any, **kwargs: Any) -> Dict[str, Any]:
        raise NotImplementedError("Excel graph unmerge_cells is not implemented.")

    async def create_spreadsheet(self, title: str) -> Dict[str, Any]:
        wb = openpyxl.Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        raw = buf.read()
        safe = title.replace(":", "_").replace("/", "_") or "Workbook"
        if not safe.lower().endswith(".xlsx"):
            safe = f"{safe}.xlsx"
        path = f"/me/drive/root:/{quote(safe)}:/content"
        resp = await self.graph_request(
            "PUT",
            path,
            content=raw,
            headers={
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            },
        )
        if resp.status_code not in (200, 201):
            raise RuntimeError(
                f"create workbook failed: {resp.status_code} {resp.text[:400]}"
            )
        meta = resp.json()
        iid = meta.get("id")
        return {
            "spreadsheetId": iid,
            "spreadsheetUrl": meta.get("webUrl"),
            "properties": {"title": title},
        }

    async def create_worksheet(
        self,
        title: str,
        spreadsheet_url_or_id: Optional[str] = None,
        rows: int = 1000,
        cols: int = 26,
    ) -> Dict[str, Any]:
        _ = rows
        _ = cols
        item_id = self._resolve_workbook(spreadsheet_url_or_id)
        async with self._workbook_session(item_id) as session_id:
            data = await self.graph_request(
                "POST",
                f"/me/drive/items/{item_id}/workbook/worksheets/add",
                json_body={"name": title},
                headers={"workbook-session-id": session_id},
            )
            if data.status_code not in (200, 201):
                raise RuntimeError(data.text[:400])
            return data.json() if data.content else {"success": True}

    async def update_worksheet(
        self,
        worksheet_title: str,
        spreadsheet_url_or_id: Optional[str] = None,
        new_title: Optional[str] = None,
        rows: Optional[int] = None,
        cols: Optional[int] = None,
        hidden: Optional[bool] = None,
        tab_color: Optional[Dict[str, float]] = None,
    ) -> Dict[str, Any]:
        if any(x is not None for x in (rows, cols, hidden, tab_color)):
            raise NotImplementedError(
                "Microsoft Excel via Graph supports only worksheet rename (new_title), "
                "not grid size / hidden / tab color."
            )
        if not new_title:
            raise ValueError(
                "Microsoft Excel: provide new_title to rename the worksheet."
            )
        item_id = self._resolve_workbook(spreadsheet_url_or_id)
        seg = self._ws_seg(worksheet_title)
        async with self._workbook_session(item_id) as session_id:
            path = f"/me/drive/items/{item_id}/workbook/{seg}"
            data = await self.graph_request(
                "PATCH",
                path,
                json_body={"name": new_title},
                headers={"workbook-session-id": session_id},
            )
            if data.status_code not in (200, 204):
                raise RuntimeError(data.text[:400])
            return {"success": True}

    async def delete_worksheet(
        self,
        worksheet_title: str,
        spreadsheet_url_or_id: Optional[str] = None,
    ) -> Dict[str, Any]:
        item_id = self._resolve_workbook(spreadsheet_url_or_id)
        seg = self._ws_seg(worksheet_title)
        async with self._workbook_session(item_id) as session_id:
            path = f"/me/drive/items/{item_id}/workbook/{seg}"
            resp = await self.graph_request(
                "DELETE",
                path,
                headers={"workbook-session-id": session_id},
            )
            if resp.status_code not in (200, 204):
                raise RuntimeError(resp.text[:400])
            return {"success": True}

    async def share_spreadsheet(
        self,
        spreadsheet_url_or_id: Optional[str] = None,
        share_type: str = "link",
        link_scope: str = "anyone",
        email: Optional[str] = None,
        role: str = "reader",
    ) -> Dict[str, Any]:
        item_id = self._resolve_workbook(spreadsheet_url_or_id)
        if share_type == "link":
            stype = "view" if role in ("reader", "read", "view") else "edit"
            scope = "anonymous" if link_scope == "anyone" else "organization"
            data = await self.graph_json(
                "POST",
                f"/me/drive/items/{item_id}/createLink",
                json_body={"type": stype, "scope": scope},
            )
            link_url = (data or {}).get("link", {}).get("webUrl")
            if link_url:
                return {"webViewLink": link_url}
            return {"success": True, "result": data}
        if share_type == "user" and not email:
            raise ValueError("email is required for user share")
        invite_role = "read" if role in ("reader", "read") else "write"
        data = await self.graph_json(
            "POST",
            f"/me/drive/items/{item_id}/invite",
            json_body={
                "recipients": [{"email": email}],
                "roles": [invite_role],
            },
        )
        return {"success": True, "result": data}

    async def delete_spreadsheet(
        self, spreadsheet_url_or_id: Optional[str] = None
    ) -> bool:
        item_id = self._resolve_workbook(spreadsheet_url_or_id)
        await self.graph_json("DELETE", f"/me/drive/items/{item_id}", ok=(204,))
        return True

    @tool(name="excel__read_spreadsheet")
    async def _t_read_spreadsheet(
        self,
        spreadsheet_url_or_id: Annotated[Optional[str], "Spreadsheet URL or ID"] = None,
        range_name: Annotated[
            Optional[str],
            "A1 notation range (e.g. 'A1:D10'). Default: empty (reads entire worksheet)",  # noqa: E501
        ] = None,
        worksheet_title: Annotated[
            Optional[str], "Worksheet title to read from"
        ] = None,
    ) -> str:
        """Read data from an Excel spreadsheet."""
        result = await self.read_spreadsheet(
            spreadsheet_url_or_id=spreadsheet_url_or_id,
            range_name=range_name if range_name is not None else "",
            worksheet_title=worksheet_title,
        )
        return json.dumps(result, indent=2)

    @tool(name="excel__update_spreadsheet")
    async def _t_update_spreadsheet(
        self,
        spreadsheet_url_or_id: Annotated[Optional[str], "Spreadsheet URL or ID"] = None,
        range_name: Annotated[
            Optional[str], "A1 notation range (e.g. 'A1:D10'). Default: empty"
        ] = None,
        values: Annotated[
            Optional[List[List[Any]]], "2D array of values to write"
        ] = None,
        value_input_option: Annotated[
            Optional[str],
            "How to interpret input values: 'RAW' or 'USER_ENTERED'. Default: 'RAW'",  # noqa: E501
        ] = None,
        worksheet_title: Annotated[Optional[str], "Worksheet title to update"] = None,
    ) -> str:
        """Update (overwrite) values in an Excel spreadsheet range."""
        result = await self.update_spreadsheet(
            spreadsheet_url_or_id=spreadsheet_url_or_id,
            range_name=range_name if range_name is not None else "",
            values=values,
            value_input_option=(
                value_input_option if value_input_option is not None else "RAW"
            ),
            worksheet_title=worksheet_title,
        )
        return json.dumps(result, indent=2)

    @tool(name="excel__append_spreadsheet")
    async def _t_append_spreadsheet(
        self,
        spreadsheet_url_or_id: Annotated[Optional[str], "Spreadsheet URL or ID"] = None,
        range_name: Annotated[
            Optional[str],
            "A1 notation range to append to (determines the table). Default: None",  # noqa: E501
        ] = None,
        values: Annotated[
            Optional[List[List[Any]]], "2D array of values to append"
        ] = None,
        value_input_option: Annotated[
            Optional[str],
            "How to interpret input values: 'RAW' or 'USER_ENTERED'. Default: 'RAW'",  # noqa: E501
        ] = None,
        worksheet_title: Annotated[
            Optional[str], "Worksheet title to append to"
        ] = None,
    ) -> str:
        """Append rows of data after the last row of data in an Excel spreadsheet."""  # noqa: E501
        result = await self.append_spreadsheet(
            spreadsheet_url_or_id=spreadsheet_url_or_id,
            range_name=range_name,
            values=values,
            value_input_option=(
                value_input_option if value_input_option is not None else "RAW"
            ),
            worksheet_title=worksheet_title,
        )
        return json.dumps(result, indent=2)

    @tool(name="excel__create_spreadsheet")
    async def _t_create_spreadsheet(
        self,
        title: Annotated[str, "Title for the new spreadsheet"],
    ) -> str:
        """Create a new Excel spreadsheet (workbook)."""
        result = await self.create_spreadsheet(title=title)
        return json.dumps(result, indent=2)

    @tool(name="excel__delete_spreadsheet")
    async def _t_delete_spreadsheet(
        self,
        spreadsheet_url_or_id: Annotated[
            Optional[str], "Spreadsheet URL or ID to delete"
        ] = None,
    ) -> str:
        """Delete an entire Excel spreadsheet (workbook)."""
        result = await self.delete_spreadsheet(
            spreadsheet_url_or_id=spreadsheet_url_or_id,
        )
        return json.dumps({"deleted": result}, indent=2)

    @tool(name="excel__create_worksheet")
    async def _t_create_worksheet(
        self,
        title: Annotated[str, "Title for the new worksheet"],
        spreadsheet_url_or_id: Annotated[
            Optional[str], "Spreadsheet URL or ID to add the worksheet to"
        ] = None,
        rows: Annotated[
            Optional[int], "Number of rows for the new worksheet. Default: 1000"
        ] = None,
        cols: Annotated[
            Optional[int], "Number of columns for the new worksheet. Default: 26"
        ] = None,
    ) -> str:
        """Create a new worksheet in an Excel spreadsheet."""
        result = await self.create_worksheet(
            title=title,
            spreadsheet_url_or_id=spreadsheet_url_or_id,
            rows=rows if rows is not None else 1000,
            cols=cols if cols is not None else 26,
        )
        return json.dumps(result, indent=2)

    @tool(name="excel__update_worksheet")
    async def _t_update_worksheet(
        self,
        worksheet_title: Annotated[str, "Current title of the worksheet to update"],
        spreadsheet_url_or_id: Annotated[Optional[str], "Spreadsheet URL or ID"] = None,
        new_title: Annotated[Optional[str], "New title for the worksheet"] = None,
        rows: Annotated[Optional[int], "New number of rows"] = None,
        cols: Annotated[Optional[int], "New number of columns"] = None,
        hidden: Annotated[
            Optional[bool], "Whether the worksheet should be hidden"
        ] = None,
        tab_color: Annotated[
            Optional[dict],
            'Tab color with RGB components 0–1, e.g. {"red": 1, "green": 0, "blue": 0}',  # noqa: E501
        ] = None,
    ) -> str:
        """Update properties of a worksheet (rename, resize, hide, color)."""
        result = await self.update_worksheet(
            worksheet_title=worksheet_title,
            spreadsheet_url_or_id=spreadsheet_url_or_id,
            new_title=new_title,
            rows=rows,
            cols=cols,
            hidden=hidden,
            tab_color=tab_color,
        )
        return json.dumps(result, indent=2)

    @tool(name="excel__delete_worksheet")
    async def _t_delete_worksheet(
        self,
        worksheet_title: Annotated[str, "Title of the worksheet to delete"],
        spreadsheet_url_or_id: Annotated[Optional[str], "Spreadsheet URL or ID"] = None,
    ) -> str:
        """Delete a worksheet from an Excel spreadsheet."""
        result = await self.delete_worksheet(
            worksheet_title=worksheet_title,
            spreadsheet_url_or_id=spreadsheet_url_or_id,
        )
        return json.dumps(result, indent=2)

    @tool(name="excel__batch_clear")
    async def _t_batch_clear(
        self,
        spreadsheet_url_or_id: Annotated[Optional[str], "Spreadsheet URL or ID"] = None,
        ranges: Annotated[
            Optional[List[str]], "List of A1 notation ranges to clear"
        ] = None,
        worksheet_title: Annotated[
            Optional[str], "Worksheet title to clear ranges from"
        ] = None,
    ) -> str:
        """Clear one or more ranges of values in an Excel spreadsheet."""
        result = await self.batch_clear(
            spreadsheet_url_or_id=spreadsheet_url_or_id,
            ranges=ranges,
            worksheet_title=worksheet_title,
        )
        return json.dumps(result, indent=2)

    @tool(name="excel__share_spreadsheet")
    async def _t_share_spreadsheet(
        self,
        spreadsheet_url_or_id: Annotated[Optional[str], "Spreadsheet URL or ID"] = None,
        share_type: Annotated[
            Optional[str], "Type of sharing: 'link' or 'email'. Default: 'link'"
        ] = None,
        link_scope: Annotated[
            Optional[str],
            "Link scope when share_type is 'link': 'anyone' or 'domain'. Default: 'anyone'",  # noqa: E501
        ] = None,
        email: Annotated[
            Optional[str], "Email address when share_type is 'email'"
        ] = None,
        role: Annotated[
            Optional[str],
            "Role to assign: 'reader', 'writer', or 'owner'. Default: 'reader'",
        ] = None,
    ) -> str:
        """Share an Excel spreadsheet via link or with a specific email."""
        result = await self.share_spreadsheet(
            spreadsheet_url_or_id=spreadsheet_url_or_id,
            share_type=share_type if share_type is not None else "link",
            link_scope=link_scope if link_scope is not None else "anyone",
            email=email,
            role=role if role is not None else "reader",
        )
        return json.dumps(result, indent=2)
